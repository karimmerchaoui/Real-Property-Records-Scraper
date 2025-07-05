import time

from selenium.webdriver.common.by import By
from openpyxl.styles import Alignment,Font
from openpyxl import Workbook


def create_sheet():
    workbook = Workbook()
    sheet = workbook.active

    # Define titles for each category
    general_infos_titles = ["Lowest Estimated Value", "Activity", "Parcel Number", "Owner","Mailing Address","Phone Number","Owner Occupied","Time Owned"]
    house_features_titles = [
    "Property Type",
    "Bedrooms*",
    "Total Baths",
    "Full Baths*",
    "Building Area (sq ft)",
    "Living Area (sq ft)*",
    "Garage (spaces)",
    "Year Built*"
]
    #6
    deed_titles = [
        "Document #", "Adjustable Rate Index", "Loan Amount (2nd TD)", "Contract Date",
        "Seller Name", "Rate Change Frequency", "Due Date", "Inter-family Transfer",
        "Buyer Mailing Address", "Construction Loan", "Title Company Name",
        "Buyer Name", "Buyer ID",
        "Buyer Vesting", "Recording Date", "Change Index", "Prepayment Penalty Rider (Term)",
        "Loan Amount",  "Interest Rate", "Sale Price", "Prepayment Rider", "Lender Name", "Seller ID", "Loan Type",
        "Max Interest Rate", "Document Type"
    ]
    #1
    mortgage_titles = [
        "Document #", "Loan Term Year", "Record Type", "Contract Date",
        "Rate Change Freq", "Due Date", "Page Number", "Borrower 2 ID", "Equity Credit Line", "Title Company Name",
        "Construction Loan", "Interest Rate (Not Less)", "Mailing Address", "Type of Financing", "Borrower Mailing Address",
        "DBA Name", "Recording Date", "Change Index", "Prepayment Penalty Rider (Term)", "Loan Amount", "Lender Type",
        "Interest Rate", "Prepayment Rider", "Borrower Name", "Book Number", "Vesting Type", "Record Type Code", "Lender Name", "Loan Type", "Borrower ID"
    ]
    distress_titles = [
        "Document #", "Current Beneficiary", "Original Loan Amount", "Contact Name", "Page #", "Source",
        "Minimum Bid Amount", "Telephone #", "Contract Date", "Attention To", "Past Due Amount", "As Of Date",
        "Recording Date", "Unpaid Balance", "Case #", "Auction City",
        "Original Beneficiary Code", "Loan Recording Date",
        "Auction Location", "Original Beneficiary Lender", "Auction Time", "Contact Address", "Auction Date", "Document Type Code"
    ]

    # Function to add titles horizontally under the merged cell
    def add_titles_horizontal(start_row, start_column, titles):
        col = start_column
        for title in titles:
            sheet.cell(row=start_row, column=col).value = title
            sheet.cell(row=start_row, column=col).alignment = Alignment(horizontal="center", vertical="center")
            col += 1


    # Start the first category "general infos" at column 2
    sheet.cell(row=1, column=2).value ="Address"
    sheet.cell(row=1, column=2).alignment = Alignment(horizontal="center", vertical="center")
    sheet.cell(row=1, column=2).font = Font(bold=True)
    general_infos_start_col = 2
    sheet.merge_cells(start_row=1, start_column=general_infos_start_col, end_row=1, end_column=general_infos_start_col + len(general_infos_titles) - 1)  # "general infos"
    sheet.cell(row=1, column=general_infos_start_col).value = "general infos"
    sheet.cell(row=1, column=general_infos_start_col).alignment = Alignment(horizontal="center", vertical="center")
    sheet.cell(row=1, column=general_infos_start_col).font = Font(bold=True)

    # Add general infos titles horizontally under the merged cell
    add_titles_horizontal(2, general_infos_start_col, general_infos_titles)

    # Start the second category "house features" after general infos
    house_features_start_col = general_infos_start_col + len(general_infos_titles)
    sheet.merge_cells(start_row=1, start_column=house_features_start_col, end_row=1, end_column=house_features_start_col + len(house_features_titles) - 1)  # "house features"
    sheet.cell(row=1, column=house_features_start_col).value = "house features"
    sheet.cell(row=1, column=house_features_start_col).alignment = Alignment(horizontal="center", vertical="center")
    sheet.cell(row=1, column=house_features_start_col).font = Font(bold=True)

    # Add house features titles horizontally under the merged cell
    add_titles_horizontal(2, house_features_start_col, house_features_titles)

    # Start the third category "deed" after house features
    deed_start_col = house_features_start_col + len(house_features_titles)
    sheet.merge_cells(start_row=1, start_column=deed_start_col, end_row=1, end_column=deed_start_col + len(deed_titles) - 1)  # "deed"
    sheet.cell(row=1, column=deed_start_col).value = "deed"
    sheet.cell(row=1, column=deed_start_col).alignment = Alignment(horizontal="center", vertical="center")
    sheet.cell(row=1, column=deed_start_col).font = Font(bold=True)

    # Add deed titles horizontally under the merged cell
    add_titles_horizontal(2, deed_start_col, deed_titles)

    # Start the fourth category "mortgage" after deed
    mortgage_start_col = deed_start_col + len(deed_titles)
    sheet.merge_cells(start_row=1, start_column=mortgage_start_col, end_row=1, end_column=mortgage_start_col + len(mortgage_titles) - 1)  # "mortgage"
    sheet.cell(row=1, column=mortgage_start_col).value = "mortgage"
    sheet.cell(row=1, column=mortgage_start_col).alignment = Alignment(horizontal="center", vertical="center")
    sheet.cell(row=1, column=mortgage_start_col).font = Font(bold=True)

    # Add mortgage titles horizontally under the merged cell
    add_titles_horizontal(2, mortgage_start_col, mortgage_titles)

    # Start the fifth category "distress" after mortgage
    distress_start_col = mortgage_start_col + len(mortgage_titles)
    sheet.merge_cells(start_row=1, start_column=distress_start_col, end_row=1, end_column=distress_start_col + len(distress_titles) - 1)  # "distress"
    sheet.cell(row=1, column=distress_start_col).value = "distress"
    sheet.cell(row=1, column=distress_start_col).alignment = Alignment(horizontal="center", vertical="center")
    sheet.cell(row=1, column=distress_start_col).font = Font(bold=True)

    # Add distress titles horizontally under the merged cell
    add_titles_horizontal(2, distress_start_col, distress_titles)

    # Save the workbook
    return workbook


def find_value_in_row(sheet, value, row_number, start_column, end_column):

    row = sheet[row_number]  # Access the specific row by its number

    # Loop through the specified column range
    for col in range(start_column - 1, end_column):
        cell = row[col]
        if cell.value == value:
            return cell.column  # Return the column number where the value is found

    return None  # Return None if the value is not found within the range


def list_to_dict(title_lists):
    # Assuming the first list contains titles and the second list contains values
    titles = title_lists[0]
    values = title_lists[1]

    # Create a dictionary by zipping titles and values
    result_dict = dict(zip(titles, values))

    return result_dict


def append_values_to_sheet(sheet, title_lists, title,r):

    # Get the headers in the first row (titles)
    headers = {}

    d=list_to_dict(title_lists)
    c=None
    for i in d:
        if title=="Deed":
            c=find_value_in_row(sheet,i,2,11,49)
        elif title=="Mortgage":
            c = find_value_in_row(sheet, i, 2, 46, 83)
        elif title=="Distressed":
            c = find_value_in_row(sheet, i, 2, 79, sheet.max_column)
        if c:
            sheet.cell(row=r, column=c).value = d[i]
            sheet.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center")


def save_table_to_excel(driver, table_xpath, title, wb,r):
    # Get the active sheet
    sheet = wb.active
    # Extract header (thead)
    header_row = driver.find_elements(By.XPATH, f"{table_xpath}/thead/tr/th")
    header_data = [header.text.strip() for header in header_row]
    # Extract body (tbody)
    body_rows = driver.find_elements(By.XPATH, f"{table_xpath}/tbody/tr")
    all_data = [header_data]  # Include header data
    for row in body_rows:
        cells = row.find_elements(By.TAG_NAME, "td")
        row_data = [cell.text.strip() for cell in cells]
        all_data.append(row_data)
    # Transpose data (swap rows and columns for horizontal placement)
    transposed_data = list(map(list, zip(*all_data)))

    if not transposed_data:
        return wb  # Return the workbook unchanged if no data

    # Find the next available column to write data horizontally

    append_values_to_sheet(sheet, transposed_data, title,r)

    return wb


import pandas as pd
def read_leads_from_excel(file_path):
    # Read the Excel file into a DataFrame
    df = pd.read_excel(file_path, header=None)  # Use `header=None` since there might not be a header row
    leads = df.iloc[:, 0].dropna().tolist()  # Get the first column, drop NaN, and convert to list
    return leads


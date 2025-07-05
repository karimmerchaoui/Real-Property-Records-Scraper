from selenium.common.exceptions import NoSuchElementException
import time
from elements import element_exists_xpath
from formatting import format_house_feature
from selenium.webdriver.common.by import By
from openpyxl.styles import PatternFill
from excel_helpers import find_value_in_row
def general_infos(browser, ev,house_features,workbook,pr):
    activity_x = '/html/body/rpr-app/rpr-layout/main/rpr-property-details/div[2]/div[5]/rpr-property-details-info-tab/div[2]/div[2]/div/rpr-sales-and-financing/rpr-chart-card/rpr-collapsible-panel/section/div/div/rpr-details-table/div/div[2]/table'
    parcel_x = '/html/body/rpr-app/rpr-layout/main/rpr-property-details/div[2]/div[5]/rpr-property-details-info-tab/div[2]/div[2]/div/rpr-property-details-two-column-details[2]/div/rpr-collapsible-panel/section/div/div/div/ul[1]/li[1]/div[2]/span'

    # Check if the Excel file exists

        # Create a new workbook and sheet
    sheet = workbook.active
    r=sheet.max_row+1
    light_green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    sheet.cell(row=r, column=1).value = pr
    sheet.cell(row=r, column=1).fill = light_green_fill
    row_labels = []
    row_values = []
    for i in house_features:
        c = find_value_in_row(sheet, i, 2, 6, 13)
        if c:
            sheet.cell(row=r, column=c).value = house_features[i]
    # Write the house features data
    # Extracting House features
    activity_text = ""
    parcel_text = ""

    # Extracting Activities
    if element_exists_xpath(browser, activity_x):
        event = browser.find_element('xpath',activity_x)
        activity_text = event.text


    # Extracting Parcel
    if element_exists_xpath(browser, parcel_x):
        parcel = browser.find_element('xpath',parcel_x)
        parcel_text = parcel.text

    # Extracting Owner
    owner_text = get_owner_name(browser)

    # Prepare lowest estimated value for appending
    try:
        lowest = ev.strip()
    except IndexError:
        lowest = ev
    if lowest=="No closed price available.":
        lowest="RPR price is not available"

    # Append extracted data to the Excel sheet horizontally
    lowest_i = find_value_in_row(sheet, "Lowest Estimated Value", 2, 1, 5)
    activity_text_i = find_value_in_row(sheet, "Activity", 2, 1, 5)
    parcel_text_i = find_value_in_row(sheet, "Parcel Number", 2, 1, 5)
    owner_text_i = find_value_in_row(sheet, "Owner", 2, 1, 5)
    sheet.cell(row=r, column=lowest_i).value = lowest
    sheet.cell(row=r, column=activity_text_i).value = activity_text
    sheet.cell(row=r, column=parcel_text_i).value = parcel_text
    sheet.cell(row=r, column=owner_text_i).value = owner_text



    return workbook,r

def get_owner_name(driver):

    try:
        li_elements = driver.find_elements('xpath', "//li[contains(@class, 'basic-fact ng-star-inserted')]")

        for li in li_elements:
            if "Owner Name" in li.text:
                owner_name_div = li.find_element('xpath', ".//div[contains(text(), 'Owner Name')]")

                sibling = owner_name_div.find_element('xpath', "following-sibling::*")

                return sibling.text.strip()
    except:
        print("Owner Not Found")
        return None

def get_houses_feature(browser):
    house_features_text = "Loading..."

    while "Loading..." in house_features_text:
        time.sleep(0.5)
        house_feature_x = "/html/body/rpr-app/rpr-layout/main/rpr-property-details/div[2]/div[5]/rpr-property-details-info-tab/div[2]/div[2]/div/rpr-property-details-facts/rpr-collapsible-panel/section/div/div/div/form/rpr-details-table/div/div[2]/table"
        if element_exists_xpath(browser, house_feature_x):
            house_features_text = browser.find_element('xpath', house_feature_x).text.replace("Name Public Facts Your Changes","")
            house_features_text = '\n'.join(house_features_text.split('\n')[1:])
            return format_house_feature(house_features_text,["Property Type","Bedrooms*","Total Baths","Full Baths*","Building Area (sq ft)","Living Area (sq ft)*","Garage (spaces)","Year Built*"])
    return


def extract_owner_facts(driver):
    data_dict = {}
    ul_x = '/html/body/rpr-app/rpr-layout/main/rpr-property-details/div[2]/div[5]/rpr-property-details-info-tab/div[2]/div[2]/div/rpr-property-details-two-column-details[3]/div/rpr-collapsible-panel/section/div/div/div'
    # Find all <ul> elements
    ul_e = driver.find_element(By.XPATH, ul_x)
    uls = ul_e.find_elements(By.XPATH, "//ul")

    # Loop through each <ul> and then each <li> inside
    for ul_index, ul in enumerate(uls, start=1):  # Start from 1 to match XPath indexing
        lis = ul.find_elements(By.TAG_NAME, "li")
        for li_index, li in enumerate(lis, start=1):
            try:
                # Construct XPath dynamically
                label_xpath = f"{ul_x}/ul[{ul_index}]/li[{li_index}]/div[1]"
                value_xpath = f"{ul_x}/ul[{ul_index}]/li[{li_index}]/div[2]/span"

                # Extract text
                label_element = driver.find_element(By.XPATH, label_xpath)
                value_element = driver.find_element(By.XPATH, value_xpath)

                label = label_element.text.strip()
                value = value_element.text.strip()
                data_dict[label] = value
                # Store in dictionary

            except NoSuchElementException:
                pass
    return data_dict


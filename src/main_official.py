"""
RPR Extractor GUI Application
A GUI application for extracting property data from RPR (Real Property Records).
"""

import os
import re
import time
import threading
import tkinter as tk
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
import concurrent.futures
from concurrent.futures import ThreadPoolExecutor, TimeoutError

import customtkinter as ctk
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import (
    NoSuchElementException,
    ElementClickInterceptedException,
    ElementNotInteractableException,
    JavascriptException,
    WebDriverException
)
from urllib3.exceptions import ReadTimeoutError, ProtocolError
from requests.exceptions import ConnectTimeout, ConnectionError, ReadTimeout
from openpyxl.styles import PatternFill

# Import custom modules
from elements import element_exists_id, element_exists_xpath, element_exists_tag, click_close_button, scroll
from excel_helpers import save_table_to_excel, read_leads_from_excel, create_sheet, find_value_in_row
from extracting import get_owner_name, extract_owner_facts
from formatting import format_house_feature


class RPRExtractorGUI:
    """Main GUI class for RPR Extractor application."""

    def __init__(self):
        self.processed_count = 0
        self.setup_gui()

    def setup_gui(self):
        """Initialize and configure the GUI components."""
        # Set CustomTkinter appearance
        ctk.set_appearance_mode("Dark")
        ctk.set_default_color_theme("blue")

        # Create main window
        self.root = ctk.CTk()
        self.root.title("RPR Extractor")
        self.root.geometry("600x450")

        # Create GUI components
        self._create_file_selection_frame()
        self._create_output_selection_frame()
        self._create_control_buttons()
        self._create_progress_indicators()
        self._create_progress_text()

    def _create_file_selection_frame(self):
        """Create the file selection frame and components."""
        self.file_frame = ctk.CTkFrame(self.root)
        self.file_frame.pack(pady=10, padx=10, fill="x")

        file_label = ctk.CTkLabel(self.file_frame, text="Select Excel File:")
        file_label.pack(side="left", padx=5, pady=5)

        self.file_entry = ctk.CTkEntry(self.file_frame, width=400)
        self.file_entry.pack(side="left", padx=5, pady=5)

        file_button = ctk.CTkButton(self.file_frame, text="Browse", command=self.select_file)
        file_button.pack(side="left", padx=5, pady=5)

    def _create_output_selection_frame(self):
        """Create the output folder selection frame and components."""
        self.output_frame = ctk.CTkFrame(self.root)
        self.output_frame.pack(pady=10, padx=10, fill="x")

        output_label = ctk.CTkLabel(self.output_frame, text="     Output Folder:")
        output_label.pack(side="left", padx=5, pady=5)

        self.output_entry = ctk.CTkEntry(self.output_frame, width=400)
        self.output_entry.pack(side="left", padx=5, pady=5)

        output_button = ctk.CTkButton(self.output_frame, text="Browse", command=self.select_output_folder)
        output_button.pack(side="left", padx=5, pady=5)

    def _create_control_buttons(self):
        """Create control buttons."""
        self.start_button = ctk.CTkButton(self.root, text="Start Processing", command=self.start_processing)
        self.start_button.pack(pady=5)

    def _create_progress_indicators(self):
        """Create progress indicators."""
        self.progress_var = tk.DoubleVar()
        self.progress_label = tk.Label(
            self.root,
            text="",
            fg="white",
            bg=self.root.cget("bg"),
            bd=0
        )
        self.progress_label.pack(pady=5)

        # Configure progress bar style
        style = ttk.Style()
        style.theme_use('alt')
        style.configure(
            "blue.Horizontal.TProgressbar",
            thickness=20,
            troughcolor="gray",
            background="#0078d4"
        )

        self.progress_bar = ttk.Progressbar(
            self.root,
            variable=self.progress_var,
            style='TProgressbar',
            maximum=100,
            length=450
        )
        self.progress_bar.pack(pady=0)

    def _create_progress_text(self):
        """Create progress text display."""
        self.progress_text = ctk.CTkTextbox(self.root, height=300, width=500, wrap="word")
        self.progress_text.pack(pady=10)

    def select_file(self):
        """Open file dialog to select Excel file."""
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if file_path:
            self.file_entry.delete(0, ctk.END)
            self.file_entry.insert(0, file_path)

    def select_output_folder(self):
        """Open folder dialog to select output directory."""
        folder_path = filedialog.askdirectory()
        if folder_path:
            self.output_entry.delete(0, tk.END)
            self.output_entry.insert(0, folder_path)

    def log_message(self, message):
        """Log a message to the progress text area."""
        self.progress_text.insert(ctk.END, f"\n {message} \n")
        self.progress_text.update()

    def start_processing(self):
        """Start the property processing in a separate thread."""
        file_path = self.file_entry.get()
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        output_folder = f"{self.output_entry.get()}/output_{timestamp}.xlsx"

        if not file_path:
            messagebox.showerror("Error", "Please select an Excel file.")
            return

        if os.path.exists(output_folder):
            try:
                os.remove(output_folder)
            except PermissionError:
                self.log_message(f"ERROR: Please close {output_folder}")
                return

        # Disable Start button to prevent duplicate processing
        self.start_button.configure(state="disabled")
        self.progress_text.update()

        # Create and start processing thread
        thread = threading.Thread(target=self.process_leads, args=(file_path, output_folder))
        thread.daemon = True
        thread.start()

        # Monitor thread completion
        self._monitor_thread(thread)

    def _monitor_thread(self, thread):
        """Monitor the processing thread and re-enable controls when done."""
        if thread.is_alive():
            self.root.after(100, lambda: self._monitor_thread(thread))
        else:
            self.start_button.configure(state="normal")
            self.log_message("FINISHED")

    def process_leads(self, file_path, output_folder):
        """Process all leads from the Excel file."""
        try:
            self.processed_count = 0
            leads = read_leads_from_excel(file_path)
            self.progress_var.set(0)
            total_leads = len(leads)
            self.progress_label.configure(text=f"{self.processed_count}/{total_leads}")

            workbook = create_sheet()
            workbook.save("init.xlsx")

            # Process leads with thread pool
            with concurrent.futures.ThreadPoolExecutor(max_workers=9) as executor:
                futures = [
                    executor.submit(
                        self.process_single_lead,
                        lead,
                        total_leads,
                        workbook,
                        output_folder
                    )
                    for lead in leads
                ]

                for future in concurrent.futures.as_completed(futures):
                    try:
                        future.result()
                    except Exception as e:
                        print(f"An error occurred: {e}")

        except Exception as e:
            print(f"An unexpected error occurred: {str(e)}")
            self.log_message(f"Unexpected error: {str(e)}")
        finally:
            # Save final workbook
            file_path = Path(output_folder)
            file_path.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(file_path)
            self.progress_var.set(100)
            self.progress_label.configure(text=f"{total_leads}/{total_leads}")

    def process_single_lead(self, property_address, total_leads, workbook, output_folder):
        """Process a single property lead."""
        # Clean up property address
        property_address = "\n".join(line for line in property_address.splitlines() if line.strip())
        property_address = property_address.strip()
        property_address = re.sub(r'[\/:*?"<>|]', '_', property_address)

        self.log_message(f"Processing {property_address}...")

        # Create browser instance
        browser = self._create_browser()

        try:
            # Login and process property
            self._login_to_rpr(browser)

            if not self._search_property(browser, property_address):
                return

            # Extract property data
            estimated_value = self._extract_estimated_value(browser)
            house_features = self._extract_house_features(browser)

            if house_features is None:
                self.log_message(f"Failed to load data for {property_address}")
                return

            # Save data to workbook
            workbook, row = self._save_general_info(browser, estimated_value, house_features, workbook,
                                                    property_address)
            workbook = self._save_property_records(browser, property_address, workbook, row)

            # Update progress
            self.processed_count += 1
            self.progress_var.set((self.processed_count / total_leads) * 100)
            self.progress_label.configure(text=f"{self.processed_count}/{total_leads}")

            # Save workbook
            file_path = Path(output_folder)
            file_path.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(file_path)

        except Exception as e:
            self.log_message(f"Error processing {property_address}: {str(e)}")
        finally:
            browser.quit()

    def _create_browser(self):
        """Create and configure Chrome browser instance."""
        options = Options()
        arguments = [
            '--start-minimized',
            '--no-sandbox',
            '--headless',
            '--disable-extensions',
            '--disable-dev-shm-usage',
            '--disable-notifications',
            '--disable-javascript',
            '--disable-infobars',
            '--disable-gpu',
            '--blink-settings=imagesEnabled=false',
            '--ignore-certificate-errors',
            '--window-size=500,500'
        ]

        for arg in arguments:
            options.add_argument(arg)

        return webdriver.Chrome(options=options)

    def _login_to_rpr(self, browser):
        """Login to RPR system."""
        browser.get('https://auth.narrpr.com/auth/sign-in')
        time.sleep(2)

        WebDriverWait(browser, 25).until(
            lambda b: b.execute_script("return document.readyState") == "complete"
        )

        # Handle potential dialog
        if element_exists_xpath(browser, '//*[@id="mat-mdc-dialog-0"]/div/div'):
            browser.get('https://auth.narrpr.com/auth/sign-in')
            time.sleep(3)

        WebDriverWait(browser, 25).until(
            lambda b: b.execute_script("return document.readyState") == "complete"
        )

        # Enter credentials (these should be moved to config or environment variables)
        email_field = browser.find_element("css selector", "#SignInEmail")
        email_field.clear()
        email_field.send_keys("EMAIL")  # TODO: Move to config

        password_field = browser.find_element("css selector", "#SignInPassword")
        password_field.clear()
        password_field.send_keys("PASSWORD")  # TODO: Move to config

        login_button = browser.find_element("css selector", "#SignInBtn")
        login_button.click()
        time.sleep(1)

    def _search_property(self, browser, property_address):
        """Search for a property in RPR."""
        # Wait for search bar to be available
        start_time = time.time()
        while not element_exists_xpath(browser,
                                       '/html/body/rpr-app/rpr-layout/main/rpr-home/div[1]/div/rpr-property-search-form/form/div/div[1]/div[2]/input'):
            if time.time() - start_time > 30:
                browser.refresh()
                time.sleep(1)
                start_time = time.time()
            time.sleep(0.5)

        search_bar = browser.find_element('xpath',
                                          '/html/body/rpr-app/rpr-layout/main/rpr-home/div[1]/div/rpr-property-search-form/form/div/div[1]/div[2]/input')

        # Handle potential intercepting elements
        while not self._try_click_element(search_bar):
            if element_exists_id(browser, 'mat-mdc-dialog-0'):
                try:
                    click_close_button(browser)
                except:
                    self.log_message(f"{property_address} does not exist")
                    return False

        # Perform search
        search_bar.clear()
        search_bar.send_keys(property_address)

        search_button = browser.find_element('xpath',
                                             "/html/body/rpr-app/rpr-layout/main/rpr-home/div[1]/div/rpr-property-search-form/form/div/div[3]/div/button")
        search_button.click()

        # Check if property was found
        time.sleep(1)
        try:
            not_found_element = browser.find_element('xpath',
                                                     '/html/body/rpr-app/rpr-layout/main/rpr-home/div[1]/div/rpr-property-search-form/form/div/div[1]/div[2]/div/div/div[1]')
            if "NO LOCATION FOUND" in not_found_element.text:
                self.log_message(f"Property Not Found: {property_address}")
                return False
        except NoSuchElementException:
            pass

        WebDriverWait(browser, 25).until(
            lambda b: b.execute_script("return document.readyState") == "complete"
        )

        return True

    def _try_click_element(self, element):
        """Try to click an element, handling potential interception."""
        try:
            element.click()
            return True
        except ElementClickInterceptedException:
            return False

    def _extract_estimated_value(self, browser):
        """Extract estimated value from property page."""
        scroll(browser)
        time.sleep(1)

        xpath_ev1 = "/html/body/rpr-app/rpr-layout/main/rpr-property-details/div[2]/div[5]/rpr-property-details-info-tab/div[1]/div[2]/rpr-property-details-summary-panel/div/div/div[4]/rpr-property-estimate-details/section/section[1]/section[1]/div/div[1]/span[1]"

        if element_exists_xpath(browser, xpath_ev1):
            return browser.find_element('xpath', xpath_ev1).text
        else:
            # Try alternative xpaths
            alternative_xpaths = [
                '/html/body/rpr-app/rpr-layout/main/rpr-property-details/div[2]/div[5]/rpr-property-details-info-tab/div[1]/div[2]/rpr-property-details-summary-panel/div/div/div[2]/section/div[2]',
                '/html/body/rpr-app/rpr-layout/main/rpr-property-details/div[2]/div[5]/rpr-property-details-info-tab/div[1]/div[2]/rpr-property-details-summary-panel/div/div/div[4]/rpr-property-estimate-details/section/section[1]/section[1]/div/div[1]'
            ]

            for xpath in alternative_xpaths:
                if element_exists_xpath(browser, xpath):
                    return browser.find_element('xpath', xpath).text

        return "No closed price available."

    def _extract_house_features(self, browser):
        """Extract house features from property page."""
        with ThreadPoolExecutor(max_workers=1) as executor:
            future = executor.submit(self._get_house_features, browser)
            try:
                return future.result(timeout=40)
            except TimeoutError:
                self.log_message("Failed to load the RPR page. Reloading...")
                return None

    def _get_house_features(self, browser):
        """Get house features from the page."""
        house_features_text = "Loading..."

        while "Loading..." in house_features_text:
            time.sleep(0.5)
            house_feature_xpath = "/html/body/rpr-app/rpr-layout/main/rpr-property-details/div[2]/div[5]/rpr-property-details-info-tab/div[2]/div[2]/div/rpr-property-details-facts/rpr-collapsible-panel/section/div/div/div/form/rpr-details-table/div/div[2]/table"

            if element_exists_xpath(browser, house_feature_xpath):
                house_features_text = browser.find_element('xpath', house_feature_xpath).text.replace(
                    "Name Public Facts Your Changes", "")
                house_features_text = '\n'.join(house_features_text.split('\n')[1:])

                features_to_extract = [
                    "Property Type", "Bedrooms*", "Total Baths", "Full Baths*",
                    "Building Area (sq ft)", "Living Area (sq ft)*", "Garage (spaces)", "Year Built*"
                ]

                return format_house_feature(house_features_text, features_to_extract)

        return {}

    def _save_general_info(self, browser, estimated_value, house_features, workbook, property_address):
        """Save general property information to Excel."""
        return general_infos(browser, estimated_value, house_features, workbook, property_address)

    def _save_property_records(self, browser, property_address, workbook, row):
        """Save property records to Excel."""
        return save_records(browser, property_address, workbook, row)

    def run(self):
        """Start the GUI application."""
        self.root.mainloop()


# Helper functions (extracted from original code)
def wait_for_json(driver, timeout=25):
    """Wait for JSON request to complete."""
    driver.execute_script("""
        window.jsonLoaded = false;
        let open = XMLHttpRequest.prototype.open;
        XMLHttpRequest.prototype.open = function() {
            if (arguments[1].includes(arguments[1])) {
                window.jsonLoaded = true;
            }
            open.apply(this, arguments);
        };
    """)

    WebDriverWait(driver, timeout).until(
        lambda d: d.execute_script("return window.jsonLoaded;")
    )


def general_infos(browser, estimated_value, house_features, workbook, property_address):
    """Extract and save general property information."""
    # XPath definitions
    activity_xpath = '/html/body/rpr-app/rpr-layout/main/rpr-property-details/div[2]/div[5]/rpr-property-details-info-tab/div[2]/div[2]/div/rpr-sales-and-financing/rpr-chart-card/rpr-collapsible-panel/section/div/div/rpr-details-table/div/div[2]/table'
    parcel_xpath = '/html/body/rpr-app/rpr-layout/main/rpr-property-details/div[2]/div[5]/rpr-property-details-info-tab/div[2]/div[2]/div/rpr-property-details-two-column-details[2]/div/rpr-collapsible-panel/section/div/div/div/ul[1]/li[1]/div[2]/span'

    sheet = workbook.active
    row = sheet.max_row + 1
    light_green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    # Set property address
    sheet.cell(row=row, column=1).value = property_address
    sheet.cell(row=row, column=1).fill = light_green_fill

    # Save house features
    for feature, value in house_features.items():
        col = find_value_in_row(sheet, feature, 2, 6, 13)
        if col:
            sheet.cell(row=row, column=col).value = value

    # Extract additional information
    activity_text = ""
    parcel_text = ""

    if element_exists_xpath(browser, activity_xpath):
        activity_text = browser.find_element('xpath', activity_xpath).text

    if element_exists_xpath(browser, parcel_xpath):
        parcel_text = browser.find_element('xpath', parcel_xpath).text

    owner_text = get_owner_name(browser)

    # Prepare estimated value
    try:
        lowest = estimated_value.strip()
    except (IndexError, AttributeError):
        lowest = estimated_value

    if lowest == "No closed price available.":
        lowest = "RPR price is not available"

    # Save extracted data
    lowest_col = find_value_in_row(sheet, "Lowest Estimated Value", 2, 1, 9)
    activity_col = find_value_in_row(sheet, "Activity", 2, 1, 9)
    parcel_col = find_value_in_row(sheet, "Parcel Number", 2, 1, 9)
    owner_col = find_value_in_row(sheet, "Owner", 2, 1, 9)

    if lowest_col:
        sheet.cell(row=row, column=lowest_col).value = lowest
    if activity_col:
        sheet.cell(row=row, column=activity_col).value = activity_text
    if parcel_col:
        sheet.cell(row=row, column=parcel_col).value = parcel_text
    if owner_col:
        sheet.cell(row=row, column=owner_col).value = owner_text

    # Extract and save owner facts
    owner_facts = extract_owner_facts(browser)
    for label, value in owner_facts.items():
        col = find_value_in_row(sheet, label, 2, 1, 11)
        if col:
            sheet.cell(row=row, column=col).value = value

    return workbook, row


def save_records(browser, property_address, workbook, row):
    """Save property records to Excel."""
    for i in range(0, 4):
        record_xpath = f'//*[@id="mat-tab-group-0-label-{i}"]'

        if element_exists_xpath(browser, record_xpath):
            element = browser.find_element('xpath', f'//*[@id="mat-tab-group-0-label-{i}"]/span[2]/span')

            if element_exists_xpath(browser, record_xpath) and element.text.strip() != 'Tax':
                title = element.text.strip()
                show_more_xpath = f'//*[@id="mat-tab-group-0-content-{i}"]/div/rpr-details-table/button'

                next_panel = browser.find_element('xpath', record_xpath)

                try:
                    next_panel.click()
                except (JavascriptException, ElementNotInteractableException):
                    return workbook

                time.sleep(0.2)

                record_history_xpath = f'//*[@id="mat-tab-group-0-content-{i}"]/div/rpr-details-table/div[1]/div[2]/table'

                if element_exists_xpath(browser, show_more_xpath):
                    show_more = browser.find_element('xpath', show_more_xpath)
                    try:
                        show_more.click()
                    except ElementNotInteractableException:
                        pass

                workbook = save_table_to_excel(browser, record_history_xpath, title, workbook, row)

    return workbook


def main():
    """Main entry point for the application."""
    app = RPRExtractorGUI()
    app.run()


if __name__ == "__main__":
    main()